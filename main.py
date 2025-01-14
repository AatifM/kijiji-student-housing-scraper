from bs4 import BeautifulSoup
from selenium import webdriver
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

def get_links(url: str) -> list:
    ad_links = []
    response = requests.get(url)

    if response.status_code != 200:
        print(f"Failed to reach page: {url}")

    else:
        soup = BeautifulSoup(response.text, "lxml")
        ads = soup.find_all("a", attrs={"data-testid": "listing-link"})

        for ad in ads:
            link = ad.get("href")
            ad_links.append(link)
    
    return ad_links



def scrape_info(ad_links: list) -> list:
    rows = []

    for advert in (ad_links):
        response = requests.get(advert)
        soup = BeautifulSoup(response.text, "lxml")

        try:
            title = soup.find("h1").text

        except AttributeError:
            title = ""

        try:
            price = soup.find("span", attrs={"itemprop": "price"}).text

        except AttributeError:
            price = ""

        try:
            address = soup.find("span", attrs={"itemprop": "address"}).text

        except AttributeError:
            address = ""

        try:
            num_bedrooms = soup.find("span", attrs={"itemprop": "numberOfBedrooms"}).text

        except AttributeError:
            num_bedrooms = ""

        try:
            num_bathrooms = soup.find("span", attrs={"itemprop": "numberOfBathroomsTotal"}).text

        except AttributeError:
            num_bathrooms = ""

        try:
            date_posted = soup.find("div", attrs={"itemprop": "datePosted"})['content']

        except (AttributeError, TypeError):
            date_posted = ""

        rows.append({
            "title": title,
            "price": price,
            "address": address,
            "num_bedrooms": num_bedrooms,
            "num_bathrooms": num_bathrooms,
            "date_posted": date_posted,
            "url": advert
            })
        
    return rows



def convert_to_xlsx(rows:list) -> None:
    df = pd.DataFrame(rows)
    excel_file = "kijiji_student_housing.xlsx"
    df.to_excel(excel_file, sheet_name="Listings", index=False)

    # Now that we have all the listings in the spreadsheet adjust it
    wb = load_workbook(excel_file)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    columns_to_wrap = ["title", "address"]
    for col in columns_to_wrap:
        col_idx = df.columns.get_loc(col) + 1
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = wrap_alignment

    url_col_idx = df.columns.get_loc("url") + 1
    for row in ws.iter_rows(min_row=2, min_col=url_col_idx, max_col=url_col_idx):
        for cell in row:
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0000FF", underline="single")

    right_alignment = Alignment(horizontal="right", vertical="center")
    columns_to_align = ["price", "num_bedrooms", "num_bathrooms", "date_posted"]
    for col in columns_to_align:
        col_idx = df.columns.get_loc(col) + 1
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = right_alignment

    ws.freeze_panes = "A2"
    wb.save(excel_file)



def main():
    url = input("Enter url: ")
    ad_links = get_links(url)
    rows = scrape_info(ad_links)
    convert_to_xlsx(rows)



if __name__ == "__main__":
    main()
